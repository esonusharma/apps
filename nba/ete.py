import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import io
import zipfile

st.set_page_config(page_title="ETE Marks Mapper", layout="wide")
st.title("📊 ETE Marks Mapper with Highlighting")

highlight_red_ids = ['2010990024', '2055991123', '2055991126', '2055991600']

# --- Function to clean columns ---
def clean_col(col):
    return col.fillna('').astype(str).str.strip().str.replace(r'\.0$', '', regex=True)

# --- Mapping logic ---
def map_marks(super_df, input_df):
    super_id_col = 'id'
    super_course_col = 'course-code'
    input_id_col = 'Admission No. (Roll No.)'
    input_course_col = 'Course Code'

    super_df[super_id_col] = clean_col(super_df[super_id_col])
    super_df[super_course_col] = clean_col(super_df[super_course_col])
    input_df[input_id_col] = clean_col(input_df[input_id_col])
    input_df[input_course_col] = clean_col(input_df[input_course_col])

    q1_parts = [
        'Obtained Marks Of Q1 \n (a)', 'Obtained Marks Of Q1 \n (b)',
        'Obtained Marks Of Q1 \n (c)', 'Obtained Marks Of Q1 \n (d)',
        'Obtained Marks Of Q1 \n (e)'
    ]
    if all(part in input_df.columns for part in q1_parts):
        input_df[q1_parts] = input_df[q1_parts].apply(pd.to_numeric, errors='coerce')
        input_df['Obtained Marks Of Q1'] = input_df[q1_parts].sum(axis=1)

    available_questions = []
    for i in range(1, 17):
        col = f'Obtained Marks Of Q{i}'
        if col in input_df.columns:
            input_df[col] = pd.to_numeric(input_df[col], errors='coerce')
            available_questions.append(i)

    input_df = input_df[[input_id_col, input_course_col] + [f'Obtained Marks Of Q{i}' for i in available_questions]]

    merged_df = pd.merge(
        super_df,
        input_df,
        how='left',
        left_on=[super_id_col, super_course_col],
        right_on=[input_id_col, input_course_col]
    )

    for i in available_questions:
        src_col = f'Obtained Marks Of Q{i}'
        dest_col = f'ete-q{i}'
        if dest_col in merged_df.columns:
            merged_df[dest_col] = merged_df[src_col]

    drop_cols = [input_id_col, input_course_col] + [f'Obtained Marks Of Q{i}' for i in available_questions]
    merged_df.drop(columns=drop_cols, inplace=True, errors='ignore')

    for i in range(1, 17):
        col = f'ete-q{i}'
        if col not in merged_df.columns:
            merged_df[col] = pd.NA

    for i in range(1, 17):
        col = f'ete-q{i}'
        if col in merged_df.columns:
            merged_df[col] = merged_df[col].replace(0, 'U')

    def highlight_rule(row):
        if all(pd.isna(row[f'ete-q{i}']) for i in range(1, 17)):
            return 'red' if row[super_id_col] in highlight_red_ids else 'yellow'
        return ''
    merged_df['__highlight__'] = merged_df.apply(highlight_rule, axis=1)

    return merged_df

# --- Export to Excel with highlights ---
def export_with_highlight(df, filename="Mapped_ETE_Marks_Highlighted.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapped Marks"

    headers = df.drop(columns='__highlight__').columns.tolist()
    for col_idx, col in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=col)

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for i, row in enumerate(df.itertuples(index=False), start=2):
        row_data = row[:-1]
        highlight = row[-1]
        for j, val in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if highlight == 'red':
                cell.fill = red_fill
            elif highlight == 'yellow':
                cell.fill = yellow_fill

    wb.save(filename)
    return filename

# --- Sidebar Upload ---
st.sidebar.header("📁 Upload Files")
super_file = st.sidebar.file_uploader("Upload main-file (Excel)", type=["xlsx"])
input_files = st.sidebar.file_uploader("Upload one or more subject-file(s)", type=["xlsx"], accept_multiple_files=True)

# --- Sidebar Sample Generator ---
if st.sidebar.button("📥 Download Sample Input Files"):
    # Sample main-file
    main_sample = pd.DataFrame({
        'sno': [1],
        'id': ['1234567890'],
        'name': ['John Doe'],
        'course-code': ['ME101'],
        'st1-marks': [10],
        'st2-marks': [12],
        'ete-marks': [35],
        **{f'A{i}': [''] for i in range(1, 5)},
        **{f'st1-{i}': [1] for i in range(1, 14)},
        **{f'st2-{i}': [2] for i in range(1, 14)},
        **{f'ete-q{i}': [None] for i in range(1, 17)}
    })

    subject_sample = pd.DataFrame({
        'Admission No. (Roll No.)': ['1234567890'],
        'Course Code': ['ME101'],
        'Obtained Marks Of Q1 \n (a)': [1],
        'Obtained Marks Of Q1 \n (b)': [2],
        'Obtained Marks Of Q1 \n (c)': [3],
        'Obtained Marks Of Q1 \n (d)': [4],
        'Obtained Marks Of Q1 \n (e)': [5],
        **{f'Obtained Marks Of Q{i}': [i] for i in range(2, 17)}
    })

    # Save both to buffers
    main_buffer = io.BytesIO()
    subject_buffer = io.BytesIO()
    main_sample.to_excel(main_buffer, index=False)
    subject_sample.to_excel(subject_buffer, index=False)
    main_buffer.seek(0)
    subject_buffer.seek(0)

    # Create ZIP file in memory
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zf:
        zf.writestr("main-file.xlsx", main_buffer.read())
        zf.writestr("subject-file.xlsx", subject_buffer.read())
    zip_buffer.seek(0)

    st.sidebar.download_button(
        label="⬇️ Download Sample Files (.zip)",
        data=zip_buffer,
        file_name="sample_input_files.zip",
        mime="application/zip"
    )

# --- Processing Logic ---
if super_file and input_files:
    try:
        super_df = pd.read_excel(super_file, dtype=str)
        input_df_list = [pd.read_excel(file, dtype=str) for file in input_files]
        input_df = pd.concat(input_df_list, ignore_index=True)

        if 'id' not in super_df.columns or 'course-code' not in super_df.columns:
            st.error("❌ main-file must contain 'id' and 'course-code'")
        elif 'Admission No. (Roll No.)' not in input_df.columns or 'Course Code' not in input_df.columns:
            st.error("❌ subject-file(s) must contain 'Admission No. (Roll No.)' and 'Course Code'")
        else:
            final_df = map_marks(super_df, input_df)

            st.subheader("✅ Mapped & Highlighted Preview")
            st.dataframe(final_df.drop(columns='__highlight__'), use_container_width=True)

            filename = "Mapped_ETE_Marks_Highlighted.xlsx"
            export_with_highlight(final_df, filename)

            with open(filename, "rb") as f:
                st.download_button("⬇️ Download Highlighted Excel", f, file_name=filename)

    except Exception as e:
        st.error(f"❌ Error occurred: {e}")
else:
    st.info("👈 Please upload the main-file and at least one subject-file.")
