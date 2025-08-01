import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import io
import zipfile
import os

st.sidebar.title(":rainbow[Dr. Sonu Sharma Apps]")
st.sidebar.subheader("Input/Output")

st.title("📊 ETE Marks Mapping")
st.header(":green[ETE]", divider="rainbow")
st.subheader(":red[Mapping of ETE question marks from subject-file(s) to main-file]", divider="rainbow")

highlight_red_ids = ['2010990024', '2055991123', '2055991126', '2055991600']

def clean_col(col):
    return col.fillna('').astype(str).str.strip().str.replace(r'\.0$', '', regex=True)

def map_marks(main_df, subject_df):
    main_id_col = 'id'
    main_course_col = 'course-code'
    subject_id_col = 'Admission No. (Roll No.)'
    subject_course_col = 'Course Code'

    main_df[main_id_col] = clean_col(main_df[main_id_col])
    main_df[main_course_col] = clean_col(main_df[main_course_col])
    subject_df[subject_id_col] = clean_col(subject_df[subject_id_col])
    subject_df[subject_course_col] = clean_col(subject_df[subject_course_col])

    q1_parts = [
        'Obtained Marks Of Q1 \n (a)', 'Obtained Marks Of Q1 \n (b)',
        'Obtained Marks Of Q1 \n (c)', 'Obtained Marks Of Q1 \n (d)',
        'Obtained Marks Of Q1 \n (e)'
    ]
    if all(part in subject_df.columns for part in q1_parts):
        subject_df[q1_parts] = subject_df[q1_parts].apply(pd.to_numeric, errors='coerce')
        subject_df['Obtained Marks Of Q1'] = subject_df[q1_parts].sum(axis=1)

    available_questions = []
    for i in range(1, 17):
        col = f'Obtained Marks Of Q{i}'
        if col in subject_df.columns:
            subject_df[col] = pd.to_numeric(subject_df[col], errors='coerce')
            available_questions.append(i)

    subject_df = subject_df[[subject_id_col, subject_course_col] + [f'Obtained Marks Of Q{i}' for i in available_questions]]

    merged_df = pd.merge(
        main_df,
        subject_df,
        how='left',
        left_on=[main_id_col, main_course_col],
        right_on=[subject_id_col, subject_course_col]
    )

    for i in available_questions:
        src_col = f'Obtained Marks Of Q{i}'
        dest_col = f'ete-q{i}'
        if dest_col in merged_df.columns:
            merged_df[dest_col] = merged_df[src_col]

    drop_cols = [subject_id_col, subject_course_col] + [f'Obtained Marks Of Q{i}' for i in available_questions]
    merged_df.drop(columns=drop_cols, inplace=True, errors='ignore')

    for i in range(1, 17):
        col = f'ete-q{i}'
        if col not in merged_df.columns:
            merged_df[col] = pd.NA

    for i in range(1, 17):
        col = f'ete-q{i}'
        if col in merged_df.columns:
            merged_df[col] = merged_df[col].replace(0, 'U')

    def highlight_rule(row):
        if all(pd.isna(row[f'ete-q{i}']) for i in range(1, 17)):
            return 'red' if row[main_id_col] in highlight_red_ids else 'yellow'
        return ''
    merged_df['__highlight__'] = merged_df.apply(highlight_rule, axis=1)

    return merged_df

def export_with_highlight(df, filename="Mapped_ETE_Marks_Highlighted.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapped Marks"

    headers = df.drop(columns='__highlight__').columns.tolist()
    for col_idx, col in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=col)

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for i, row in enumerate(df.itertuples(index=False), start=2):
        row_data = row[:-1]
        highlight = row[-1]
        for j, val in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if highlight == 'red':
                cell.fill = red_fill
            elif highlight == 'yellow':
                cell.fill = yellow_fill

    wb.save(filename)
    return filename

main_sample = pd.DataFrame({
    'sno': [1],
    'id': ['1234567890'],
    'name': ['John Doe'],
    'course-code': ['ME101'],
    'st1-marks': [10],
    'st2-marks': [12],
    'ete-marks': [35],
    **{f'A{i}': [''] for i in range(1, 5)},
    **{f'st1-{i}': [1] for i in range(1, 14)},
    **{f'st2-{i}': [2] for i in range(1, 14)},
    **{f'ete-q{i}': [None] for i in range(1, 17)}
})

subject_sample = pd.DataFrame({
    'Admission No. (Roll No.)': ['1234567890'],
    'Course Code': ['ME101'],
    'Obtained Marks Of Q1 \n (a)': [1],
    'Obtained Marks Of Q1 \n (b)': [2],
    'Obtained Marks Of Q1 \n (c)': [3],
    'Obtained Marks Of Q1 \n (d)': [4],
    'Obtained Marks Of Q1 \n (e)': [5],
    **{f'Obtained Marks Of Q{i}': [i] for i in range(2, 17)}
})

# Save to memory buffers
main_buffer = io.BytesIO()
subject_buffer = io.BytesIO()
main_sample.to_excel(main_buffer, index=False)
subject_sample.to_excel(subject_buffer, index=False)
main_buffer.seek(0)
subject_buffer.seek(0)

# Create zip archive in memory
zip_buffer = io.BytesIO()
with zipfile.ZipFile(zip_buffer, 'w') as zf:
    zf.writestr("main-file.xlsx", main_buffer.read())
    zf.writestr("subject-file.xlsx", subject_buffer.read())
zip_buffer.seek(0)

st.sidebar.download_button(
    label="⬇️ Download Sample Input Files (.zip)",
    data=zip_buffer,
    file_name="sample_input_files.zip",
    mime="application/zip"
)

# ---------------- Sidebar Input ----------------
st.sidebar.header("📂 Upload Files")
main_file = st.sidebar.file_uploader("Upload Main File", type=["xlsx"])
subject_files = st.sidebar.file_uploader("Upload Subject File(s)", type=["xlsx"], accept_multiple_files=True)

# ---------------- Mapping and Output ----------------
if main_file and subject_files:
    try:
        main_df = pd.read_excel(main_file, dtype=str)
        subject_df_list = [pd.read_excel(file, dtype=str) for file in subject_files]
        subject_df = pd.concat(subject_df_list, ignore_index=True)

        if 'id' not in main_df.columns or 'course-code' not in main_df.columns:
            st.error("❌ Main file must contain 'id' and 'course-code'.")
        elif 'Admission No. (Roll No.)' not in subject_df.columns or 'Course Code' not in subject_df.columns:
            st.error("❌ Subject file(s) must contain 'Admission No. (Roll No.)' and 'Course Code'.")
        else:
            final_df = map_marks(main_df, subject_df)

            st.subheader("✅ Mapped Excel Main File Preview")
            st.dataframe(final_df.drop(columns='__highlight__'), use_container_width=True)

            original_name = os.path.splitext(main_file.name)[0]
            filename = f"output_{original_name}.xlsx"
            export_with_highlight(final_df, filename)

            with open(filename, "rb") as f:
                st.download_button("⬇️ Download Mapped Excel Main File", f, file_name=filename)

    except Exception as e:
        st.error(f"❌ Error occurred: {e}")
else:
    st.info("👆 Please upload both the Main File and at least one Subject File.")
