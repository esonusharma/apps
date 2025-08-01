import streamlit as st
import pandas as pd
import numpy as np
import io
import random
import zipfile
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

st.sidebar.title(":rainbow[Dr. Sonu Sharma Apps]")
st.sidebar.subheader("Input/Output")

sample_data = pd.DataFrame({
    'sno': [1, 2, 3],
    'id': ['2410994001', '2410994002', '2410994003'],
    'name': ['avik', 'sanya', 'aman'],
    'course-code': ['24ME0101', '24ME0102', '24ME0103'],
    'st1-marks': [39, 38, 36],
    'st2-marks': [38, 39, 37],
    'ete-marks': [58, 58, 55]
})
sample_buffer = io.BytesIO()
sample_data.to_excel(sample_buffer, index=False)
sample_buffer.seek(0)
st.sidebar.download_button(
    label="📅 Download Sample Input File",
    data=sample_buffer,
    file_name="sample_input.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

uploaded_files = st.sidebar.file_uploader("Upload Excel files", type=["xlsx"], accept_multiple_files=True)
process_button = st.sidebar.button("Start Processing")

st.title("📊 Drawing Marks Processing Panel")
st.header(":green[ST1, ST2, ETE]", divider="rainbow")
st.subheader(":red[Divides obtained marks of ST1, ST2, and ETE of Drawing subjects in questions with no choices]", divider="rainbow")

# Structures
structure_st = {
    1: 2, 2: 2, 3: 2, 4: 2, 5: 2,
    6: 5, 7: 5, 8: 5, 9: 5,
    10: 10
}

structure_ete = {
    1: 2, 2: 2, 3: 2, 4: 2, 5: 2, 6: 2, 7: 2, 8: 2, 9: 2, 10: 2,
    11: 5, 12: 5, 13: 5, 14: 5,
    15: 10, 16: 10
}

# Split marks function without U

def split_marks_no_U(total, structure):
    keys = list(structure.keys())
    max_values = [structure[k] for k in keys]
    while True:
        assigned = [random.randint(0, mx) for mx in max_values]
        raw_total = sum(assigned)
        if raw_total == 0:
            continue
        scaled = [int(val * total / raw_total) for val in assigned]
        for i in range(len(scaled)):
            scaled[i] = min(scaled[i], max_values[i])
        diff = total - sum(scaled)
        attempts = 0
        while diff != 0 and attempts < 1000:
            for i in range(len(scaled)):
                if diff == 0:
                    break
                if diff > 0 and scaled[i] < max_values[i]:
                    scaled[i] += 1
                    diff -= 1
                elif diff < 0 and scaled[i] > 0:
                    scaled[i] -= 1
                    diff += 1
            attempts += 1
        if sum(scaled) == total:
            break
    return dict(zip(keys, scaled))

# Excel styling function
def style_excel(file_buffer, highlight_rows):
    wb = load_workbook(file_buffer)
    ws = wb.active

    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal="center", vertical="center")
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center
            cell.border = thin
            if cell.row == 1:
                cell.fill = yellow
            elif cell.row - 2 in highlight_rows:
                cell.fill = red

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    styled_io = io.BytesIO()
    wb.save(styled_io)
    styled_io.seek(0)
    return styled_io

# Row processing function
def process_row(row):
    invalid = False
    out = row.to_dict()
    mapping = {
        'st1-marks': ('st1', 50, structure_st),
        'st2-marks': ('st2', 50, structure_st),
        'ete-marks': ('ete', 60, structure_ete)
    }

    for col_name, (prefix, max_val, struct) in mapping.items():
        val = row.get(col_name)
        try:
            val = int(val)
        except:
            invalid = True
            continue
        if val > max_val or val < 0:
            invalid = True
        else:
            split = split_marks_no_U(val, struct)
            for k, v in split.items():
                out[f"{prefix}-q{k}"] = v
    return out, invalid

# File processing function
def process_file(file):
    df = pd.read_excel(file)
    processed = []
    invalid_rows = []

    for idx, row in df.iterrows():
        processed_row, is_invalid = process_row(row)
        processed.append(processed_row)
        if is_invalid:
            invalid_rows.append(idx)

    out_df = pd.DataFrame(processed)
    for col in out_df.columns:
        out_df[col] = out_df[col].apply(lambda x: x if np.isscalar(x) else str(x))
    buffer = io.BytesIO()
    out_df.to_excel(buffer, index=False)
    buffer.seek(0)
    styled_output = style_excel(buffer, invalid_rows)
    return styled_output

# Handle processing and zip download
if process_button and uploaded_files:
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zipf:
        for file in uploaded_files:
            styled_output = process_file(file)
            zipf.writestr(f"processed_{file.name}", styled_output.read())
    zip_buffer.seek(0)
    st.success("✅ Processing complete! Download your ZIP below.")
    st.download_button(
        label="📆 Download All Processed Files as ZIP",
        data=zip_buffer,
        file_name="all_processed_files.zip",
        mime="application/zip"
    )
