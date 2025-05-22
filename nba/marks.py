import streamlit as st
import pandas as pd
import numpy as np
import io
import random
import zipfile
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Streamlit UI setup
st.sidebar.title("Input/Output")
generate_sample_button = st.sidebar.button("Generate Sample Input File")
uploaded_files = st.sidebar.file_uploader("Upload Excel files", type=["xlsx"], accept_multiple_files=True)
process_button = st.sidebar.button("Start Processing")
st.sidebar.subheader(":rainbow[Developed by Dr. Sonu Sharma]")

st.title("📊 Welcome to NBA Marks Processing Panel")
st.header("Initiative by :green[Centre of Excellence, Modelling and Simulation]", divider="rainbow")
st.subheader("Department of Mechanical Engineering, CUIET-AE, Chitkara University, Punjab", divider="rainbow")

# Structures
structure_13 = {
    1: 5,
    2: 2, 3: 2, 4: 2, 5: 2, 6: 2, 7: 2,
    8: 5, 9: 5, 10: 5, 11: 5,
    12: 10, 13: 10
}
structure_16 = {
    'q1': 5,
    'q2': 2, 'q3': 2, 'q4': 2, 'q5': 2, 'q6': 2, 'q7': 2,
    'q8': 5, 'q9': 5, 'q10': 5, 'q11': 5, 'q12': 5, 'q13': 5,
    'q14': 10, 'q15': 10, 'q16': 10
}
na_groups_13 = [[2, 3, 4, 5, 6, 7], [8, 9, 10, 11], [12, 13]]
na_groups_16 = [['q2', 'q3', 'q4', 'q5', 'q6', 'q7'], ['q8', 'q9', 'q10', 'q11', 'q12', 'q13'], ['q14', 'q15', 'q16']]

# Sample file generation
if generate_sample_button:
    sample_data = pd.DataFrame({
        'sno': [1, 2, 3],
        'id': ['2410994001', '2410994002', '2410994003'],
        'name': ['avik', 'sanya', 'aman'],
        'course-code': ['24ME0101', '24ME0102', '24ME0103'],
        'st1-marks': [39, 38, 36],
        'st2-marks': [38, 39, 37],
        'ete-marks': [58, 58, 55]
    })
    sample_buffer = io.BytesIO()
    sample_data.to_excel(sample_buffer, index=False)
    sample_buffer.seek(0)
    st.sidebar.download_button(
        label="📥 Download Sample Input File",
        data=sample_buffer,
        file_name="sample_input.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Split marks function
def split_marks(total, structure, na_groups):
    max_cols = list(structure.keys())
    result = {}
    na_choices = [random.choice(group) for group in na_groups]
    allowed_cols = [col for col in max_cols if col not in na_choices]
    max_values = [structure[col] for col in allowed_cols]

    while True:
        assigned = [random.randint(0, mx) for mx in max_values]
        raw_total = sum(assigned)
        if raw_total == 0:
            continue
        scaled = [int(val * total / raw_total) for val in assigned]
        for i, col in enumerate(allowed_cols):
            scaled[i] = min(scaled[i], structure[col])
        diff = total - sum(scaled)
        attempts = 0
        while diff != 0 and attempts < 1000:
            for i, col in enumerate(allowed_cols):
                if diff == 0:
                    break
                if diff > 0 and scaled[i] < structure[col]:
                    scaled[i] += 1
                    diff -= 1
                elif diff < 0 and scaled[i] > 0:
                    scaled[i] -= 1
                    diff += 1
            attempts += 1
        if sum(scaled) == total:
            break

    idx = 0
    for col in max_cols:
        if col in na_choices:
            result[col] = "N/A"
        else:
            result[col] = scaled[idx]
            idx += 1
    return result

# Excel styling function
def style_excel(file_buffer, highlight_rows):
    wb = load_workbook(file_buffer)
    ws = wb.active

    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal="center", vertical="center")
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center
            cell.border = thin
            if cell.row == 1:
                cell.fill = yellow
            elif cell.row - 2 in highlight_rows:
                cell.fill = red

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    styled_io = io.BytesIO()
    wb.save(styled_io)
    styled_io.seek(0)
    return styled_io

# Row processing function
def process_row(row):
    invalid = False
    out = row.to_dict()
    mapping = {'st1-marks': ('st1', 40, structure_13, na_groups_13),
               'st2-marks': ('st2', 40, structure_13, na_groups_13),
               'ete-marks': ('ete', 60, structure_16, na_groups_16)}

    for col_name, (prefix, max_val, struct, na) in mapping.items():
        val = row.get(col_name)
        try:
            val = int(val)
        except:
            invalid = True
            continue
        if val > max_val or val < 0:
            invalid = True
        else:
            split = split_marks(val, struct, na)
            for k, v in split.items():
                out[f"{prefix}-{k}"] = v
    return out, invalid

# File processing

def process_file(file):
    df = pd.read_excel(file)
    processed = []
    invalid_rows = []

    for idx, row in df.iterrows():
        processed_row, is_invalid = process_row(row)
        processed.append(processed_row)
        if is_invalid:
            invalid_rows.append(idx)

    out_df = pd.DataFrame(processed)
    for col in out_df.columns:
        out_df[col] = out_df[col].apply(lambda x: x if np.isscalar(x) else str(x))
    buffer = io.BytesIO()
    out_df.to_excel(buffer, index=False)
    buffer.seek(0)
    styled_output = style_excel(buffer, invalid_rows)
    return styled_output

# Handle processing and zip download
if process_button and uploaded_files:
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zipf:
        for file in uploaded_files:
            styled_output = process_file(file)
            zipf.writestr(f"processed_{file.name}", styled_output.read())
    zip_buffer.seek(0)
    st.success("✅ Processing complete! Download your ZIP below.")
    st.download_button(
        label="📦 Download All Processed Files as ZIP",
        data=zip_buffer,
        file_name="all_processed_files.zip",
        mime="application/zip"
    )