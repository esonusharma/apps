import streamlit as st
import pandas as pd
import numpy as np
import io
import random
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

st.title("ST Marks Splitter")

st.markdown("[Sample Input File](nba-st.streamlit.app/static/sample-input-file.xlsx)")

uploaded_files = st.file_uploader("Upload Excel files", type=["xlsx"], accept_multiple_files=True)

def distribute_marks(total_marks):
    structure = {
        1: 5,
        2: 2, 3: 2, 4: 2, 5: 2, 6: 2, 7: 2,
        8: 5, 9: 5, 10: 5, 11: 5,
        12: 10, 13: 10
    }

    na_col_2_7 = random.choice(range(2, 8))
    na_col_8_11 = random.choice(range(8, 12))
    na_col_12_13 = random.choice(range(12, 14))
    na_columns = {na_col_2_7, na_col_8_11, na_col_12_13}

    result = {}
    allowed_cols = [col for col in range(1, 14) if col not in na_columns]
    max_values = [structure[col] for col in allowed_cols]

    while True:
        assigned = [random.randint(0, mx) for mx in max_values]
        raw_total = sum(assigned)
        if raw_total == 0:
            continue

        # Scale to match total_marks
        scaled = [int(val * total_marks / raw_total) for val in assigned]

        # Clamp to max values
        for i, col in enumerate(allowed_cols):
            scaled[i] = min(scaled[i], structure[col])

        # Adjust to fix rounding error
        current_sum = sum(scaled)
        diff = total_marks - current_sum
        attempts = 0

        while diff != 0 and attempts < 1000:
            for i, col in enumerate(allowed_cols):
                if diff == 0:
                    break
                if diff > 0 and scaled[i] < structure[col]:
                    scaled[i] += 1
                    diff -= 1
                elif diff < 0 and scaled[i] > 0:
                    scaled[i] -= 1
                    diff += 1
            attempts += 1

        if sum(scaled) == total_marks:
            break

    idx = 0
    for col in range(1, 14):
        if col in na_columns:
            result[str(col)] = "N/A"
        else:
            result[str(col)] = scaled[idx]
            idx += 1

    return result

def style_excel(file_buffer):
    wb = load_workbook(file_buffer)
    ws = wb.active

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
            if cell.row == 1:
                cell.fill = header_fill

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    styled_output = io.BytesIO()
    wb.save(styled_output)
    styled_output.seek(0)
    return styled_output

if uploaded_files:
    for uploaded_file in uploaded_files:
        df = pd.read_excel(uploaded_file)

        if 'marks' not in df.columns:
            st.error(f"'marks' column not found in {uploaded_file.name}")
            continue

        split_rows = []
        for _, row in df.iterrows():
            total = int(row['marks']) if not pd.isna(row['marks']) else 0
            split = distribute_marks(total)
            split_rows.append(split)

        split_df = pd.DataFrame(split_rows)
        final_df = pd.concat([df, split_df], axis=1)

        buffer = io.BytesIO()
        final_df.to_excel(buffer, index=False)
        buffer.seek(0)

        styled_excel = style_excel(buffer)

        st.download_button(
            label=f"Download Processed: {uploaded_file.name}",
            data=styled_excel,
            file_name=f"output_{uploaded_file.name}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
