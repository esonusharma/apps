import streamlit as st
import pandas as pd
import numpy as np
import io
import random
import zipfile
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

st.sidebar.title(":rainbow[Dr. Sonu Sharma Apps]")
st.sidebar.subheader("Input/Output")

sample_data = pd.DataFrame({'marks': [39.5, 38, 36.5, 40, 21]})
sample_buffer = io.BytesIO()
sample_data.to_excel(sample_buffer, index=False)
sample_buffer.seek(0)
st.sidebar.download_button(
    label="📥 Download Sample Input File",
    data=sample_buffer,
    file_name="sample_input.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

uploaded_files = st.sidebar.file_uploader("Upload Excel files (1 column: marks)", type=["xlsx"], accept_multiple_files=True)

num_divisions = st.sidebar.number_input("Number of divisions", min_value=1, max_value=100, value=5, step=1)
division_type = st.sidebar.selectbox("Division type", ["Equal", "Random"])

max_per_component = st.sidebar.number_input(
    "Max marks per component (optional)", min_value=0.0, value=100.0, step=0.5,
    help="Only applies to Equal division. Set to 0 to disable limit."
)

process_button = st.sidebar.button("Start Processing")

st.title("📊 Marks Division Panel")
st.header(":green[Divide 'marks' Column into Divisions]")
st.subheader(":blue[Supports Equal and Random Divisions with Decimal/Integer Precision]")

def split_single_column_marks(total, divisions, division_type, max_component):
    total = round(float(total), 2)

    if division_type == "Equal":
        base = round(total / divisions, 2)
        result = [base] * divisions
        diff = round(total - sum(result), 2)
        result[-1] = round(result[-1] + diff, 2)

        # Apply max per component if set
        if max_component > 0:
            adjusted = []
            for val in result:
                adjusted.append(min(val, max_component))
            sum_adj = sum(adjusted)
            if sum_adj < total:
                remaining = round(total - sum_adj, 2)
                for i in range(len(adjusted)):
                    if adjusted[i] < max_component:
                        addable = min(max_component - adjusted[i], remaining)
                        adjusted[i] = round(adjusted[i] + addable, 2)
                        remaining = round(remaining - addable, 2)
                        if remaining <= 0:
                            break
            result = adjusted

    else:  # Random
        if total == 0:
            return [0] * divisions

        int_part = int(total)
        frac_part = round(total - int_part, 2)

        if frac_part == 0:
            # Integer-only distribution
            if int_part < divisions:
                parts = [1]*int_part + [0]*(divisions - int_part)
                random.shuffle(parts)
            else:
                points = sorted(random.sample(range(1, int_part + divisions), divisions - 1))
                parts = [points[0]]
                for i in range(1, len(points)):
                    parts.append(points[i] - points[i-1])
                parts.append(int_part + divisions - points[-1])
                parts = [p - 1 for p in parts]
            result = parts
        else:
            # Distribute int_part first
            if int_part < divisions:
                parts = [1]*int_part + [0]*(divisions - int_part)
                random.shuffle(parts)
            else:
                points = sorted(random.sample(range(1, int_part + divisions), divisions - 1))
                parts = [points[0]]
                for i in range(1, len(points)):
                    parts.append(points[i] - points[i-1])
                parts.append(int_part + divisions - points[-1])
                parts = [p - 1 for p in parts]

            # Add fraction to one part
            idx = random.randint(0, divisions - 1)
            parts[idx] = round(parts[idx] + frac_part, 2)
            result = parts

    # Final correction to sum
    result = [round(max(0, val), 2) for val in result]
    diff = round(total - sum(result), 2)
    if abs(diff) > 0.01:
        result[-1] = round(result[-1] + diff, 2)
    return result

def style_excel(file_buffer, highlight_rows):
    wb = load_workbook(file_buffer)
    ws = wb.active

    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal="center", vertical="center")
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center
            cell.border = thin
            if cell.row == 1:
                cell.fill = yellow
            elif cell.row - 2 in highlight_rows:
                cell.fill = red

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    styled_io = io.BytesIO()
    wb.save(styled_io)
    styled_io.seek(0)
    return styled_io

def process_row(row):
    out = {}
    invalid = False
    val = row.get('marks')
    try:
        val = float(val)
        if val < 0:
            invalid = True
    except:
        invalid = True
        val = 0

    divisions_values = split_single_column_marks(val, num_divisions, division_type, max_per_component)
    for i, v in enumerate(divisions_values, start=1):
        out[f"div_{i}"] = v

    out['marks'] = val
    return out, invalid

def process_file(file):
    df = pd.read_excel(file)
    if 'marks' not in df.columns:
        st.warning(f"File '{file.name}' does not contain 'marks' column.")
        return None
    processed = []
    invalid_rows = []

    for idx, row in df.iterrows():
        processed_row, is_invalid = process_row(row)
        processed.append(processed_row)
        if is_invalid:
            invalid_rows.append(idx)

    out_df = pd.DataFrame(processed)
    buffer = io.BytesIO()
    out_df.to_excel(buffer, index=False)
    buffer.seek(0)
    styled_output = style_excel(buffer, invalid_rows)
    return styled_output

if process_button and uploaded_files:
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zipf:
        for file in uploaded_files:
            styled_output = process_file(file)
            if styled_output is not None:
                zipf.writestr(f"processed_{file.name}", styled_output.read())
    zip_buffer.seek(0)
    st.success("✅ Processing complete! Download your ZIP below.")
    st.download_button(
        label="📦 Download All Processed Files as ZIP",
        data=zip_buffer,
        file_name="all_processed_files.zip",
        mime="application/zip"
    )
