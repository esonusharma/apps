import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

st.title("📊 ETE Marks Mapper with Highlighted Output")

highlight_red_ids = ['2010990024', '2055991123', '2055991126', '2055991600']

def map_marks(super_df, input_df):
    super_id_col = 'id'
    super_course_col = 'course-code'
    input_id_col = 'Admission No. (Roll No.)'
    input_course_col = 'Course Code'

    # Combine Q1 subparts into one column
    q1_parts = [
        'Obtained Marks Of Q1 \n (a)', 'Obtained Marks Of Q1 \n (b)',
        'Obtained Marks Of Q1 \n (c)', 'Obtained Marks Of Q1 \n (d)',
        'Obtained Marks Of Q1 \n (e)'
    ]
    if all(part in input_df.columns for part in q1_parts):
        input_df['Obtained Marks Of Q1'] = input_df[q1_parts].sum(axis=1)

    # Identify available questions
    available_questions = []
    for i in range(1, 17):
        col = f'Obtained Marks Of Q{i}'
        if col in input_df.columns:
            available_questions.append(i)

    # Keep necessary columns
    mapping_columns = [input_id_col, input_course_col] + [f'Obtained Marks Of Q{i}' for i in available_questions]
    input_df = input_df[mapping_columns]

    # Merge with super_df
    merged_df = pd.merge(
        super_df,
        input_df,
        how='left',
        left_on=[super_id_col, super_course_col],
        right_on=[input_id_col, input_course_col]
    )

    # Copy marks into corresponding ete-q columns
    for i in available_questions:
        obtained_col = f'Obtained Marks Of Q{i}'
        ete_col = f'ete-q{i}'
        if ete_col in merged_df.columns:
            merged_df[ete_col] = merged_df[obtained_col]

    # Clean up temporary columns
    drop_cols = [f'Obtained Marks Of Q{i}' for i in available_questions]
    drop_cols += [input_id_col, input_course_col]
    merged_df.drop(columns=drop_cols, inplace=True, errors='ignore')

    # Add empty ete-q columns if missing
    for i in range(1, 17):
        ete_col = f'ete-q{i}'
        if ete_col not in merged_df.columns:
            merged_df[ete_col] = pd.NA

    # Replace 0 with 'U'
    for i in range(1, 17):
        col = f'ete-q{i}'
        if col in merged_df.columns:
            merged_df[col] = merged_df[col].replace(0, 'U')

    # Add highlight column
    def get_highlight(row):
        if all(pd.isna(row[f'ete-q{i}']) for i in range(1, 17)):
            return 'red' if str(row[super_id_col]) in highlight_red_ids else 'yellow'
        return ''
    merged_df['__highlight__'] = merged_df.apply(get_highlight, axis=1)

    return merged_df

def export_with_highlight(df, filename="Mapped_ETE_Marks_Highlighted.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ETE Marks"

    # Write headers
    headers = df.drop(columns='__highlight__').columns.tolist()
    for col_idx, col in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=col)

    # Fill definitions
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # Write data with highlight
    for i, row in enumerate(df.itertuples(index=False), start=2):
        row_data = row[:-1]  # data without highlight
        highlight = row[-1]  # last field is __highlight__

        for j, val in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if highlight == 'red':
                cell.fill = red_fill
            elif highlight == 'yellow':
                cell.fill = yellow_fill

    wb.save(filename)
    return filename

# Upload files
super_file = st.file_uploader("Upload Super File (Excel)", type=["xlsx"])
input_files = st.file_uploader("Upload One or More Input Files", type=["xlsx"], accept_multiple_files=True)

if super_file and input_files:
    super_df = pd.read_excel(super_file)
    input_df = pd.concat([pd.read_excel(file) for file in input_files], ignore_index=True)

    # Check column requirements
    if 'id' not in super_df.columns or 'course-code' not in super_df.columns:
        st.error("Super file must contain 'id' and 'course-code'.")
    elif 'Admission No. (Roll No.)' not in input_df.columns or 'Course Code' not in input_df.columns:
        st.error("Each input file must contain 'Admission No. (Roll No.)' and 'Course Code'.")
    else:
        mapped_df = map_marks(super_df, input_df)

        # Preview
        st.subheader("✅ Mapped Output (Preview)")
        st.dataframe(mapped_df.drop(columns='__highlight__'), use_container_width=True)

        # Generate downloadable Excel
        output_filename = "Mapped_ETE_Marks_Highlighted.xlsx"
        export_with_highlight(mapped_df, filename=output_filename)

        with open(output_filename, "rb") as f:
            st.download_button("📥 Download Highlighted Excel", f, file_name=output_filename)
else:
    st.info("📂 Upload a Super File and one or more Input Files to begin.")
